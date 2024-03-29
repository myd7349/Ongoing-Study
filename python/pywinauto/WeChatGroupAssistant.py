# -*- coding: utf-8 -*-
# yapf -i WeChatGroupAssistant.py
# .\WeChatGroupAssistant.py -g 猫言猫语 "地下交通站" "eBooks 分享"

import argparse
import collections
import os.path
import sys

import openpyxl
import psutil
import pywinauto

try:
    import win32gui
    win32gui_available = True
except (ModuleNotFoundError, ImportError):
    win32gui_available = False

GroupMember = collections.namedtuple('GroupMember', ['name', 'nickname'])


def find_pid(process_name):
    pid = -1

    for process in psutil.process_iter():
        try:
            pinfo = process.as_dict(attrs=['pid', 'name'])
        except psutil.NoSuchProcess:
            pass
        else:
            if process_name == pinfo['name']:
                pid = pinfo['pid']
                break

    return pid


def bring_to_foreground_win32gui(window_class, title):
    hwnd = win32gui.FindWindow(window_class, title)
    win32gui.SetForegroundWindow(hwnd)
    win32gui.ShowWindow(hwnd, 9)


def bring_to_foreground(application):
    top_window = application.top_window()

    top_window.minimize()
    top_window.restore()


def get_wechat_v1():
    application = pywinauto.application.Application(backend='uia')
    try:
        wechat_hwnd = pywinauto.findwindows.find_window(
            class_name='WeChatMainWndForPC', title_re=r'微信')
    except pywinauto.WindowNotFoundError as e:
        sys.exit(e)

    application.connect(handle=wechat_hwnd)

    return application['微信']


def get_wechat_v2():
    application = pywinauto.application.Application(backend='uia').connect(
        class_name='WeChatMainWndForPC',
        title='微信',
        visible_only=True,
        timeout=100)
    return application['微信']


def get_wechat_v3():
    wechat_pid = find_pid('WeChat.exe')
    if wechat_pid == -1:
        print('WeChat is not running.')
        sys.exit(0)

    wechat_application = pywinauto.application.Application(
        backend='uia').connect(process=wechat_pid)
    return wechat_application['微信']


def rectangle_center(rectangle):
    return ((rectangle.left + rectangle.right) // 2,
            (rectangle.top + rectangle.bottom) // 2)


def click(window, mouse_button='left', outline=True):
    if outline:
        window.draw_outline()

    center = rectangle_center(window.rectangle())
    pywinauto.mouse.click(button=mouse_button, coords=center)


def click_button(window, title, mouse_button='left', outline=True):
    button = window.child_window(title=title, control_type='Button')
    click(button, mouse_button, outline)


def pairwise(iterable):
    "s -> (s0, s1), (s2, s3), (s4, s5), ..."
    a = iter(iterable)
    return zip(a, a)


def get_group_members(wechat, group_name):
    click_button(wechat, '聊天')
    chat_list = wechat.child_window(title='会话', control_type='List')
    chat_list.draw_outline()

    try:
        group = chat_list.child_window(title_re=group_name + "*",
                                       control_type='ListItem')
    except pywinauto.MatchError:
        print('Failed to find group "{0}".'.format(group_name))
        raise

    click(group)

    group_chat_button = wechat \
        .children(control_type='Pane')[1] \
        .children(control_type='Pane')[1] \
        .children(control_type='Pane')[2] \
        .children(control_type='Pane')[0] \
        .children(control_type='Pane')[0] \
        .children(control_type='Pane')[0] \
        .children(control_type='Pane')[0] \
        .children(control_type='Pane')[0] \
        .children(control_type='Pane')[0] \
        .children(control_type='Pane')[0] \
        .children(control_type='Pane')[0] \
        .children(control_type='Pane')[0] \
        .children(control_type='Button')[0]
    click(group_chat_button)

    # TODO:
    # If there is no more room to expand the chat information window,
    # a SessionChatRoomDetailWnd will appear instead.
    group_information_window = wechat.child_window(title='聊天信息',
                                                   control_type='Window')
    group_information_window.draw_outline()

    try:
        click_button(group_information_window, '查看更多')
    except pywinauto.findwindows.ElementNotFoundError as e:
        print(e)

    member_list = group_information_window.child_window(title='聊天成员',
                                                        control_type='List')

    members = []
    count = 0

    for member in pairwise(member_list.descendants(control_type='Button')):
        name, nickname = member[0].texts()[0], member[1].texts()[0]
        count += 1

        print(f'{count:03} Name: {name}, Nickname: {nickname}')
        if not name and nickname in ('添加', '移出'):
            continue

        group_member = GroupMember(name, nickname)
        members.append(group_member)

    return members


def is_member_in_group(member, group_members, match_by='both'):
    # TODO: It is too slow.
    names = [m.name for m in group_members]
    nicknames = [m.nickname for m in group_members]

    if match_by == 'both':
        result = member in group_members
        if result:
            return str(result)
        else:
            return '{0}({1},{2})'.format(result, member.name in names,
                                         member.nickname in nicknames)
    elif match_by == 'name':
        return str(member.name in names)
    elif match_by == 'nickname':
        return str(member.nickname in nicknames)


def clear_worksheet(worksheet):
    worksheet.delete_rows(worksheet.min_row,
                          worksheet.max_row - worksheet.min_row + 1)


def save_to_worksheet(worksheet, members):
    clear_worksheet(worksheet)

    for member in members:
        worksheet.append(tuple(member))


def save_summary(worksheet, reference_group, group_members_dict, match_by):
    assert reference_group in group_members_dict

    clear_worksheet(worksheet)

    members = group_members_dict[reference_group]
    group_members_dict.pop(reference_group)

    other_groups = group_members_dict.keys()

    header = GroupMember._fields + tuple(
        ('In {0}?'.format(group) for group in other_groups))
    worksheet.append(header)

    for member in members:
        row = tuple(member) + tuple(
            (is_member_in_group(member, group_members_dict[group], match_by)
             for group in other_groups))
        worksheet.append(row)


def parse_args():
    prog = os.path.splitext(os.path.basename(sys.argv[0]))[0]

    parser = argparse.ArgumentParser(
        prog=prog,
        fromfile_prefix_chars='@',
        description='Get a list of members of a WeChat group.')

    parser.add_argument('-g',
                        '--groups',
                        nargs='+',
                        required=True,
                        help='Specify a list of group names.')
    parser.add_argument('-s',
                        '--summary',
                        action='store_true',
                        help='Create summary report.')
    parser.add_argument('-m',
                        '--match-by',
                        choices=['name', 'nickname', 'both'],
                        default='both',
                        help='Match strategy.')

    return parser.parse_args()


def main():
    args = parse_args()

    wechat = get_wechat_v2()
    #wechat.Properties.print_control_identifiers()

    if win32gui_available:
        bring_to_foreground_win32gui('WeChatMainWndForPC', '微信')

    group_members_dict = {}

    for group in args.groups:
        if group in group_members_dict:
            print('Group "{0}" already handled.'.format(group))
            continue

        members = get_group_members(wechat, group)
        group_members_dict[group] = members

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        save_to_worksheet(worksheet, members)

        workbook.save('{0}.xlsx'.format(group))
        workbook.close()

    if len(args.groups) > 1 and args.summary:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        save_summary(worksheet, args.groups[0], group_members_dict,
                     args.match_by)

        workbook.save('NotOneLess.xlsx')
        workbook.close()


if __name__ == '__main__':
    main()

# References:
# [Python pywinauto search windows with partial title](https://stackoverflow.com/questions/28216222/python-pywinauto-search-windows-with-partial-title)
# [Using PyWinAuto to control a currently running application](https://stackoverflow.com/questions/39021888/using-pywinauto-to-control-a-currently-running-application)
# [Python Automatically Collects WeChat Contacts](https://programmer.help/blogs/python-automatically-collects-wechat-contacts.html)
# [Pywinauto how do I get the list of returned elements](https://stackoverflow.com/questions/46432544/pywinauto-how-do-i-get-the-list-of-returned-elements)
# [How to access the control identifiers in pywinauto](https://stackoverflow.com/questions/5039642/how-to-access-the-control-identifiers-in-pywinauto)
# [Iterating over every two elements in a list [duplicate]](https://stackoverflow.com/questions/5389507/iterating-over-every-two-elements-in-a-list)
# [Pywinauto: unable to bring window to foreground](https://stackoverflow.com/questions/39794729/pywinauto-unable-to-bring-window-to-foreground)
# [How can I remove a key from a Python dictionary?](https://stackoverflow.com/questions/11277432/how-can-i-remove-a-key-from-a-python-dictionary)
# [How do I make a window (already running task) visible using pywinauto?](https://stackoverflow.com/questions/50332830/how-do-i-make-a-window-already-running-task-visible-using-pywinauto)
#
# PS C:\Users\myd\Desktop\Ongoing-Study\python\pywinauto\build> .\WeChatGroupAssistant.exe
# PS C:\Users\myd\Desktop\Ongoing-Study\python\pywinauto\build> .\WeChatGroupAssistant.exe -g 七小
# Traceback (most recent call last):
#   File "WeChatGroupAssistant.py", line 223, in <module>
#   File "WeChatGroupAssistant.py", line 201, in main
#   File "WeChatGroupAssistant.py", line 113, in get_group_members
# IndexError: list index out of range
# [10360] Failed to execute script WeChatGroupAssistant
# PS C:\Users\myd\Desktop\Ongoing-Study\python\pywinauto\build> .\WeChatGroupAssistant.exe -g 七小
# Traceback (most recent call last):
#   File "pywinauto\application.py", line 258, in __resolve_control
#   File "pywinauto\timings.py", line 458, in wait_until_passes
# pywinauto.timings.TimeoutError
#
# During handling of the above exception, another exception occurred:
#
# Traceback (most recent call last):
#   File "WeChatGroupAssistant.py", line 223, in <module>
#   File "WeChatGroupAssistant.py", line 201, in main
#   File "WeChatGroupAssistant.py", line 128, in get_group_members
#   File "pywinauto\application.py", line 379, in __getattribute__
#   File "pywinauto\application.py", line 261, in __resolve_control
#   File "pywinauto\timings.py", line 436, in wait_until_passes
#   File "pywinauto\application.py", line 222, in __get_ctrl
#   File "pywinauto\findwindows.py", line 87, in find_element
# pywinauto.findwindows.ElementNotFoundError: {'title': '聊天信息', 'control_type': 'Window', 'top_level_only': False, 'parent': <uia_element_info.UIAElementInfo - '微信', WeChatMainWndForPC, 2688358>, 'backend': 'uia'}
# [6476] Failed to execute script WeChatGroupAssistant
#
# [How to pad zeroes to a string?](https://stackoverflow.com/questions/339007/how-to-pad-zeroes-to-a-string)
